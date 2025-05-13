import os
import sys
import tkinter as tk
from tkinter import ttk
from mca import Region
from openpyxl import Workbook
from multiprocessing import Pool, Manager, active_children
import threading
import time
import json
from concurrent.futures import ProcessPoolExecutor, as_completed
import psutil

# ========== 用户可配置参数 ==========
# Y轴高度范围（调整方块统计的垂直范围）
Y_START = 160
Y_END = 240

# 内存优化配置（控制程序最大内存占用百分比）
MAX_MEMORY_USAGE_PERCENT = 100  # 建议值：60-80

# MCA文件目录（指向Minecraft世界的region文件夹）
MCA_DIRECTORY = r"D:\Games\Minecraft\游戏主体\.minecraft\saves\1_21_5大模型\region"

# 进程管理配置
MAX_WORKERS = None  # 最大工作进程数，None表示自动（建议保持默认）
# ==================================

# 启用高DPI支持（Windows系统）
def enable_high_dpi_support():
    if sys.platform.startswith('win'):
        from ctypes import windll
        try:
            # 尝试启用系统DPI感知
            windll.shcore.SetProcessDpiAwareness(2)  # 2 = PROCESS_PER_MONITOR_DPI_AWARE
        except:
            # 回退到系统DPI感知
            windll.user32.SetProcessDPIAware()


def get_available_memory():
    """获取可用内存百分比"""
    return psutil.virtual_memory().percent


def process_chunk(args):
    region_path, chunk_x, chunk_z = args
    block_count = {}
    try:
        region = Region.from_file(region_path)
        chunk = region.get_chunk(chunk_x, chunk_z)
        if chunk:
            for y in range(Y_START, Y_END):
                for z in range(16):
                    for x in range(16):
                        block = chunk.get_block(x, y, z)
                        if block:
                            block_id = block.id
                            block_count[block_id] = block_count.get(block_id, 0) + 1
    except Exception as e:
        print(f"处理区块 ({chunk_x},{chunk_z}) 出错: {e}", file=sys.stderr)
    return block_count


def process_mca_file(file_path, stop_event):
    if stop_event.is_set():
        return {}

    try:
        # 使用全局变量跟踪当前执行器
        global current_executor
        with ProcessPoolExecutor() as executor:
            current_executor = executor
            args_list = [(file_path, chunk_x, chunk_z) for chunk_x in range(32) for chunk_z in range(32)]
            futures = [executor.submit(process_chunk, args) for args in args_list]

            block_count = {}
            for future in as_completed(futures):
                if stop_event.is_set():
                    return block_count  # 返回已处理的部分结果
                result = future.result()
                for block_id, count in result.items():
                    block_count[block_id] = block_count.get(block_id, 0) + count
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}", file=sys.stderr)
    finally:
        current_executor = None

    return block_count


def save_progress(processed_files, total_block_count):
    try:
        with open('progress.json', 'w') as f:
            json.dump({
                "processed_files": processed_files,
                "block_count": total_block_count
            }, f)
    except Exception as e:
        print(f"保存进度时出错: {e}", file=sys.stderr)


def load_progress():
    if os.path.exists('progress.json'):
        try:
            with open('progress.json', 'r') as f:
                data = json.load(f)
                return data.get("processed_files", []), data.get("block_count", {})
        except Exception as e:
            print(f"加载进度时出错: {e}", file=sys.stderr)
    return [], {}


def terminate_all_processes():
    """终止所有子进程"""
    for p in active_children():
        try:
            p.terminate()
            p.join(timeout=1)
        except Exception as e:
            print(f"终止进程 {p.pid} 失败: {e}")

    # 尝试关闭当前执行器
    global current_executor
    if current_executor:
        try:
            current_executor.shutdown(wait=False, cancel_futures=True)
        except Exception as e:
            print(f"关闭执行器失败: {e}")
        finally:
            current_executor = None


def count_blocks_in_all_mca_files(directory, progress_bar, percentage_label, close_button, stop_event):
    mca_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.mca')]
    total_files = len(mca_files)
    if not mca_files:
        print("没有找到 .mca 文件")
        return

    processed_files, total_block_count = load_progress()
    processed_set = set(processed_files)

    progress = len(processed_files) / total_files * 100
    progress_bar['value'] = progress
    percentage_label.config(text=f"{progress:.2f}%  剩余时间: 计算中")

    start_time = time.time()
    remaining_files = [f for f in mca_files if f not in processed_set]
    total_remaining = len(remaining_files)

    print(f"开始处理 {total_remaining} 个文件...")

    # 根据系统内存动态设置最大工作进程数
    if MAX_WORKERS is None:
        max_workers = max(2, min(os.cpu_count(), 16))  # 默认2-16个进程
    else:
        max_workers = MAX_WORKERS

    print(f"使用 {max_workers} 个工作进程")

    for i, file_path in enumerate(remaining_files):
        # 监控内存使用情况
        while get_available_memory() > MAX_MEMORY_USAGE_PERCENT:
            if stop_event.is_set():
                print("用户中止处理")
                return
            time.sleep(1)  # 等待内存释放

        if stop_event.is_set():
            print("用户中止处理")
            return

        file_block_count = process_mca_file(file_path, stop_event)

        if stop_event.is_set():
            print("用户中止处理")
            return

        for block_id, count in file_block_count.items():
            total_block_count[block_id] = total_block_count.get(block_id, 0) + count

        processed_files.append(file_path)
        save_progress(processed_files, total_block_count)

        current_progress = len(processed_files) / total_files * 100
        elapsed = time.time() - start_time

        if i > 0:
            avg_time = elapsed / (i + 1)
            remaining_time = avg_time * (total_remaining - i - 1)
            remaining_str = time.strftime("%H:%M:%S", time.gmtime(remaining_time))
        else:
            remaining_str = "计算中"

        # 使用局部函数确保参数正确传递
        def update_progress(p=current_progress, s=remaining_str, c=i + 1, t=total_remaining):
            progress_bar['value'] = p
            percentage_label.config(text=f"{p:.2f}%  剩余时间: {s} ({c}/{t})")

        root.after(0, update_progress)

    if os.path.exists('progress.json'):
        os.remove('progress.json')

    excel_filename = f"{Y_START}_{Y_END}.xlsx"
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '方块id'
    ws['B1'] = '对应方块数量'

    for row in [(block_id, count) for block_id, count in total_block_count.items()]:
        ws.append(row)

    wb.save(excel_filename)
    print(f"所有文件的统计结果已保存到 {excel_filename} 中。")
    print(f"共处理 {len(processed_files)} 个文件，发现 {len(total_block_count)} 种不同方块。")

    # 完成后更新UI
    def finish_ui():
        if progress_bar.winfo_exists():
            progress_bar.destroy()
        percentage_label.config(text=f"已完成！共处理 {total_files} 个文件")
        close_button.config(text="关闭")

    root.after(0, finish_ui)


def start_processing(directory):
    global root, current_executor
    current_executor = None
    root = tk.Tk()
    root.title("处理 .mca 文件进度")
    root.geometry("400x110")  # 窗口宽度增加100像素
    root.resizable(False, False)

    # 创建居中的信息标签
    info_frame = tk.Frame(root)
    info_frame.pack(fill=tk.X, padx=10, pady=3)  # 保持原始间距

    # 使用系统默认字体
    percentage_label = tk.Label(info_frame, text="0.00%  剩余时间: 计算中")
    percentage_label.pack(expand=True)

    # 创建进度条（长度增加100像素）
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack(pady=3)  # 保持原始间距

    # 创建停止/关闭按钮
    stop_event = threading.Event()

    def on_close():
        if progress_bar.winfo_exists():
            stop_event.set()
            percentage_label.config(text="正在停止...")

            # 立即终止所有子进程
            root.after(500, terminate_all_processes)

            # 关闭窗口
            root.after(1000, root.destroy)
        else:
            root.destroy()

    # 创建按钮容器
    button_frame = tk.Frame(root)
    button_frame.pack(pady=1)  # 保持原始间距

    # 使用系统默认字体
    close_button = tk.Button(button_frame, text="停止", command=on_close)
    close_button.pack()

    # 启动处理线程
    threading.Thread(target=count_blocks_in_all_mca_files,
                     args=(directory, progress_bar, percentage_label, close_button, stop_event),
                     daemon=True).start()

    root.mainloop()


if __name__ == "__main__":
    # 启用高DPI支持
    enable_high_dpi_support()

    # 使用前面定义的MCA目录
    start_processing(MCA_DIRECTORY)
