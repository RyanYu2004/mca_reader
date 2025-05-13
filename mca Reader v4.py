import os
import sys
import tkinter as tk
from tkinter import ttk
from mca import Region
from openpyxl import Workbook
from multiprocessing import Pool
import threading
import time
import json


# 定义 y 轴范围（可根据需求修改，此处固定为 Minecraft 新版本范围）
Y_START = 160
Y_END = 240


# 定义一个函数来处理单个块
def process_chunk(args):
    region, chunk_x, chunk_z = args
    block_count = {}
    try:
        chunk = region.get_chunk(chunk_x, chunk_z)
        if chunk:
            for y in range(Y_START, Y_END):
                for z in range(16):
                    for x in range(16):
                        try:
                            block = chunk.get_block(x, y, z)
                            block_id = block.id
                            if block_id in block_count:
                                block_count[block_id] += 1
                            else:
                                block_count[block_id] = 1
                        except Exception as block_err:
                            if str(block_err) != 'Tag data does not exist':
                                print(f"在块 ({chunk_x}, {chunk_z}) 中获取方块 ({x}, {y}, {z}) 时出错: {block_err}",
                                      file=sys.stderr)
    except Exception as chunk_err:
        print(f"获取块 ({chunk_x}, {chunk_z}) 时出错: {chunk_err}", file=sys.stderr)
    return block_count


# 定义一个函数来处理单个 .mca 文件
def process_mca_file(file_path):
    try:
        # 加载 .mca 文件
        region = Region.from_file(file_path)

        # 生成所有块的参数列表
        chunk_args = [(region, chunk_x, chunk_z) for chunk_x in range(32) for chunk_z in range(32)]

        # 使用多进程池并行处理块
        with Pool() as pool:
            results = pool.map(process_chunk, chunk_args)

        # 合并所有块的统计结果
        block_count = {}
        for result in results:
            for block_id, count in result.items():
                if block_id in block_count:
                    block_count[block_id] += count
                else:
                    block_count[block_id] = 1

        return block_count
    except FileNotFoundError:
        print(f"错误: 文件 {file_path} 未找到。", file=sys.stderr)
        return {}
    except Exception as e:
        print(f"处理文件 {file_path} 时发生未知错误: {e}", file=sys.stderr)
        return {}


def save_progress(processed_files, total_block_count):
    progress_data = {
        "processed_files": processed_files,
        "block_count": total_block_count
    }
    with open('progress.json', 'w') as f:
        json.dump(progress_data, f)


def load_progress():
    if os.path.exists('progress.json'):
        with open('progress.json', 'r') as f:
            progress_data = json.load(f)
            return progress_data["processed_files"], progress_data["block_count"]
    return [], {}


def count_blocks_in_all_mca_files(directory, progress_bar, percentage_label):
    # 获取目录下所有 .mca 文件
    mca_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.mca')]
    total_files = len(mca_files)
    processed_files, total_block_count = load_progress()

    # 计算已完成的进度
    progress = len(processed_files) / total_files * 100
    progress_bar['value'] = progress

    if progress > 0:
        elapsed_time = 0  # 这里简单假设为 0，实际可以记录更准确的时间
        if len(processed_files) > 0:
            avg_time_per_file = elapsed_time / len(processed_files)
            remaining_files = total_files - len(processed_files)
            remaining_time = avg_time_per_file * remaining_files
            remaining_time_str = time.strftime("%H:%M:%S", time.gmtime(remaining_time))
        else:
            remaining_time_str = "N/A"
        percentage_label.config(text=f"{progress:.2f}%  剩余时间: {remaining_time_str}")

    start_time = time.time()
    for index, file_path in enumerate(mca_files):
        if file_path in processed_files:
            continue
        file_block_count = process_mca_file(file_path)
        # 合并当前文件的统计结果到总统计结果中
        for block_id, count in file_block_count.items():
            if block_id in total_block_count:
                total_block_count[block_id] += count
            else:
                total_block_count[block_id] = 1

        processed_files.append(file_path)
        save_progress(processed_files, total_block_count)

        # 更新进度条和百分比标签以及剩余时间
        progress = len(processed_files) / total_files * 100
        elapsed_time = time.time() - start_time
        if index > 0:
            avg_time_per_file = elapsed_time / len(processed_files)
            remaining_files = total_files - len(processed_files)
            remaining_time = avg_time_per_file * remaining_files
            remaining_time_str = time.strftime("%H:%M:%S", time.gmtime(remaining_time))
        else:
            remaining_time_str = "N/A"

        progress_bar['value'] = progress
        percentage_label.config(text=f"{progress:.2f}%  剩余时间: {remaining_time_str}")
        progress_bar.update()

    # 删除进度保存文件
    if os.path.exists('progress.json'):
        os.remove('progress.json')

    # 创建一个新的 Excel 工作簿，文件名根据 y 轴范围生成
    excel_filename = f"{Y_START}_{Y_END}.xlsx"
    wb = Workbook()
    ws = wb.active

    # 设置表头
    ws['A1'] = '方块id'
    ws['B1'] = '对应方块数量'

    # 批量写入数据
    data = [(block_id, count) for block_id, count in total_block_count.items()]
    ws.append(['方块id', '对应方块数量'])
    for row in data:
        ws.append(row)

    # 保存 Excel 文件
    wb.save(excel_filename)
    print(f"所有文件的统计结果已保存到 {excel_filename} 中。")


def start_processing(directory):
    root = tk.Tk()
    root.title("处理 .mca 文件进度")

    # 创建百分比和剩余时间标签
    percentage_label = tk.Label(root, text="0.00%  剩余时间: N/A")
    percentage_label.pack(pady=10)

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress_bar.pack(pady=20)

    # 在新线程中运行处理函数，避免阻塞 GUI
    processing_thread = threading.Thread(target=count_blocks_in_all_mca_files,
                                         args=(directory, progress_bar, percentage_label))
    processing_thread.start()

    root.mainloop()


if __name__ == "__main__":
    # 替换为你的 .mca 文件所在的目录路径
    mca_directory = r"D:\Games\Minecraft\游戏主体\.minecraft\saves\1_21_5大模型\region"
    start_processing(mca_directory)